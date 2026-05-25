"""
Servicio de Exportación a Excel
"""
from typing import List, Optional
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.config.settings import settings
from app.utils import logger, format_datetime


class ExportService:
    """Servicio para exportar datos a Excel"""
    
    @staticmethod
    def export_inspections(
        inspections_data: List[dict],
        filename: Optional[str] = None
    ) -> Optional[str]:
        """
        Exportar inspecciones a Excel
        
        Args:
            inspections_data: Lista de inspecciones en formato dict
            filename: Nombre del archivo (opcional)
            
        Returns:
            str: Ruta del archivo generado o None si error
        """
        try:
            if not inspections_data:
                logger.warning("No hay datos para exportar")
                return None
            
            # Generar nombre de archivo si no se proporciona
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"inspecciones_{timestamp}.xlsx"
            
            # Ruta completa
            filepath = Path(settings.EXPORT_PATH) / filename
            
            # Crear DataFrame
            df = pd.DataFrame(inspections_data)
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Inspecciones"
            
            # Estilos
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Agregar datos
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Estilo de encabezado
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            wb.save(filepath)
            
            logger.info(f"Archivo Excel generado: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {e}")
            return None
    
    @staticmethod
    def export_clients(clients_data: List[dict], filename: Optional[str] = None) -> Optional[str]:
        """
        Exportar clientes a Excel
        
        Args:
            clients_data: Lista de clientes
            filename: Nombre del archivo
            
        Returns:
            str: Ruta del archivo generado
        """
        try:
            if not clients_data:
                return None
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"clientes_{timestamp}.xlsx"
            
            filepath = Path(settings.EXPORT_PATH) / filename
            
            # Crear DataFrame y exportar
            df = pd.DataFrame(clients_data)
            df.to_excel(filepath, index=False, sheet_name='Clientes')
            
            logger.info(f"Clientes exportados: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error al exportar clientes: {e}")
            return None
    
    @staticmethod
    def export_equipment(equipment_data: List[dict], filename: Optional[str] = None) -> Optional[str]:
        """
        Exportar equipos a Excel
        
        Args:
            equipment_data: Lista de equipos
            filename: Nombre del archivo
            
        Returns:
            str: Ruta del archivo generado
        """
        try:
            if not equipment_data:
                return None
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"equipos_{timestamp}.xlsx"
            
            filepath = Path(settings.EXPORT_PATH) / filename
            
            df = pd.DataFrame(equipment_data)
            df.to_excel(filepath, index=False, sheet_name='Equipos')
            
            logger.info(f"Equipos exportados: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error al exportar equipos: {e}")
            return None
    
    @staticmethod
    def export_parameters_report(
        inspection_id: int,
        inspection_data: dict,
        parameters_data: List[dict],
        filename: Optional[str] = None
    ) -> Optional[str]:
        """
        Exportar reporte de parámetros de una inspección
        
        Args:
            inspection_id: ID de la inspección
            inspection_data: Datos de la inspección
            parameters_data: Parámetros medidos
            filename: Nombre del archivo
            
        Returns:
            str: Ruta del archivo generado
        """
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"reporte_inspeccion_{inspection_id}_{timestamp}.xlsx"
            
            filepath = Path(settings.EXPORT_PATH) / filename
            
            wb = Workbook()
            
            # Hoja 1: Información de la inspección
            ws1 = wb.active
            ws1.title = "Información"
            
            ws1['A1'] = "REPORTE DE INSPECCIÓN"
            ws1['A1'].font = Font(size=16, bold=True)
            
            row = 3
            for key, value in inspection_data.items():
                ws1[f'A{row}'] = str(key).replace('_', ' ').title()
                ws1[f'B{row}'] = str(value)
                ws1[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Hoja 2: Parámetros
            ws2 = wb.create_sheet("Parámetros")
            if parameters_data:
                df = pd.DataFrame(parameters_data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws2.append(r)
            
            wb.save(filepath)
            
            logger.info(f"Reporte de inspección generado: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error al generar reporte: {e}")
            return None
